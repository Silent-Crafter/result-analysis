from openpyxl import Workbook
from parser import parse
from openpyxl.chart import PieChart, Reference
from pandas import DataFrame


class Spreadsheet:
    def __init__(self, file_name, data):
        self.file_name = file_name
        self.data = data

    def create(self):
        workbook = Workbook()
        workbook.create_sheet("ANALYSIS")

        df = DataFrame(self.data["STUDENT_INFO"]).T
        counter = 0
        for student in self.data["STUDENT_INFO"].keys():
            workbook.create_sheet(student)
            worksheet = workbook[student]
            student_data = self.data["STUDENT_INFO"][student]
            marks_data = self.data["RESULT"]

            worksheet.cell(1, 1, "Seat No:")
            worksheet.cell(1, 2, student)

            worksheet.cell(2, 1, "Student Name:")
            worksheet.cell(2, 2, student_data["NAME"])

            worksheet.cell(2, 4, "Mother Name:")
            worksheet.cell(2, 5, student_data["MOTHER"])

            worksheet.cell(4, 1, "Sem")
            worksheet.cell(4, 2, "SubCode")
            worksheet.cell(4, 3, "Subject Name")
            worksheet.cell(4, 4, "Crd")
            worksheet.cell(4, 5, "Grd")
            worksheet.cell(4, 6, "GP")

            cell_index_row = 5
            while True:
                try:
                    if marks_data[counter]["SEAT NO"] != student:
                        break

                    curr_sem = marks_data[counter]["SEM"]
                except (KeyError, IndexError):
                    break
                worksheet.cell(cell_index_row, 1, curr_sem)

                try:
                    while marks_data[counter]["SEM"] == curr_sem:
                        if marks_data[counter]["SEAT NO"] != student:
                            break
                        cell_index_col = 2
                        worksheet.cell(cell_index_row, cell_index_col, marks_data[counter]["SUBJECT"])
                        worksheet.cell(cell_index_row, cell_index_col + 1,
                                       self.data["SUBJECTS"][marks_data[counter]["SUBJECT"]])
                        worksheet.cell(cell_index_row, cell_index_col + 2, marks_data[counter]["CR"])
                        worksheet.cell(cell_index_row, cell_index_col + 3, marks_data[counter]["GRADE"])
                        worksheet.cell(cell_index_row, cell_index_col + 4, marks_data[counter]["GP"])
                        cell_index_row += 1
                        counter += 1
                except IndexError:
                    break

        worksheet = workbook["ANALYSIS"]

        passed = len(df.loc[df["CGPA" if self.data["EXAM"] == "SEM2" else "SGPA"] > 0])
        failed = len(df.loc[df["CGPA" if self.data["EXAM"] == "SEM2" else "SGPA"] == -1])

        worksheet.append(["PASSED", passed])
        worksheet.append(["FAILED", failed])

        piechart = PieChart()
        labels = Reference(worksheet, min_row=1, min_col=1, max_row=2)
        data = Reference(worksheet, min_row=1, min_col=2, max_row=2)

        piechart.add_data(data)
        piechart.set_categories(labels)

        worksheet.add_chart(piechart, 'C1')

        workbook.save('demo.xlsx')
        workbook.close()


if __name__ == '__main__':
    data = parse('endsem gazette.pdf')
    wb = Spreadsheet('demo.xlsx', data)
    wb.create()
    print(data["RESULT"])
